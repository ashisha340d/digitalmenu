import base64
import hashlib
import json
import os
import re
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
IMAGES_DIR = os.path.join(ROOT, "images")
LAYOUT_PATH = os.path.join(ROOT, "layout.json")
WORKBOOK_CANDIDATES = ["Menu.xlsx", "Menu.xls", "menu.xlsx", "menu.xls"]

WS_MAGIC = "258EAFA5-E914-47DA-95CA-C5AB0DC85B11"
ws_clients = []
ws_clients_lock = threading.Lock()

# Set by /api/shutdown so the main thread (blocked in shutdown_event.wait(), below)
# can stop the server cleanly from outside — e.g. the tray launcher on Exit/Restart —
# without relying on console signals, which a windowless child process can't receive.
shutdown_event = threading.Event()


def ws_accept_key(key):
    return base64.b64encode(hashlib.sha1((key + WS_MAGIC).encode("utf-8")).digest()).decode("ascii")


def ws_frame(text):
    payload = text.encode("utf-8")
    length = len(payload)
    if length <= 125:
        header = bytes([0x81, length])
    elif length <= 65535:
        header = bytes([0x81, 126]) + length.to_bytes(2, "big")
    else:
        header = bytes([0x81, 127]) + length.to_bytes(8, "big")
    return header + payload


def broadcast_ws(message):
    frame = ws_frame(message)
    with ws_clients_lock:
        dead = []
        for sock in ws_clients:
            try:
                sock.sendall(frame)
            except Exception:
                dead.append(sock)
        for sock in dead:
            ws_clients.remove(sock)


def find_workbook_path():
    for name in WORKBOOK_CANDIDATES:
        path = os.path.join(ROOT, name)
        if os.path.exists(path):
            return path
    return os.path.join(ROOT, "Menu.xlsx")


def normalize_header(value):
    return (value or "").strip().lower().replace(" ", "").replace("_", "")


def slugify(text):
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def unique_image_filename(base_slug, ext=".webp"):
    candidate = base_slug + ext
    n = 2
    while os.path.exists(os.path.join(IMAGES_DIR, candidate)):
        candidate = f"{base_slug}-{n}{ext}"
        n += 1
    return candidate


def read_menu_items(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append((cell.value or "").strip())

    index_map = {}
    for idx, header in enumerate(headers):
        index_map[normalize_header(header)] = idx

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        item = {
            "category": row[index_map.get("category", 0)] if index_map.get("category") is not None and index_map.get("category") < len(row) else "",
            "category_hindi": row[index_map.get("categoryhindi", 0)] if index_map.get("categoryhindi") is not None and index_map.get("categoryhindi") < len(row) else "",
            "name": row[index_map.get("name", 0)] if index_map.get("name") is not None and index_map.get("name") < len(row) else "",
            "name_hindi": row[index_map.get("namehindi", 0)] if index_map.get("namehindi") is not None and index_map.get("namehindi") < len(row) else "",
            "price": row[index_map.get("price", 0)] if index_map.get("price") is not None and index_map.get("price") < len(row) else "",
            "available": str(row[index_map.get("available", 0)] if index_map.get("available") is not None and index_map.get("available") < len(row) else "true").strip().lower() not in {"false", "0", "no", "n"},
            "is_morning": str(row[index_map.get("ismorning", 0)] if index_map.get("ismorning") is not None and index_map.get("ismorning") < len(row) else "false").strip().lower() in {"true", "1", "yes", "y"},
            "image": row[index_map.get("image", 0)] if index_map.get("image") is not None and index_map.get("image") < len(row) else "",
        }
        rows.append(item)
    return rows


def write_menu_items(path, items):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = ["Category", "Category_Hindi", "Name", "Name_Hindi", "Price", "Available", "IsMorning", "Image"]
    ws.delete_rows(2, ws.max_row)
    ws.delete_cols(len(headers) + 1, ws.max_column - len(headers))
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("category", ""))
        ws.cell(row=row_idx, column=2, value=item.get("category_hindi", ""))
        ws.cell(row=row_idx, column=3, value=item.get("name", ""))
        ws.cell(row=row_idx, column=4, value=item.get("name_hindi", ""))
        ws.cell(row=row_idx, column=5, value=item.get("price", ""))
        ws.cell(row=row_idx, column=6, value="true" if item.get("available", True) else "false")
        ws.cell(row=row_idx, column=7, value="true" if item.get("is_morning", False) else "false")
        ws.cell(row=row_idx, column=8, value=item.get("image", ""))
    wb.save(path)


class MenuHandler(BaseHTTPRequestHandler):
    def _send_json(self, payload, status=200):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_file(self, relative_path):
        if relative_path in {"", "/"}:
            relative_path = "/admin.html"
        safe_path = relative_path.lstrip("/")
        if safe_path.startswith("api/"):
            return None
        full_path = os.path.normpath(os.path.join(ROOT, safe_path))
        if not full_path.startswith(ROOT):
            self.send_error(403)
            return None
        if not os.path.exists(full_path):
            self.send_error(404)
            return None
        ext = os.path.splitext(full_path)[1].lower()
        mimetype = "text/html; charset=utf-8" if ext in {".html", ".htm"} else "application/octet-stream"
        if ext in {".js", ".mjs"}:
            mimetype = "application/javascript; charset=utf-8"
        elif ext in {".css"}:
            mimetype = "text/css; charset=utf-8"
        elif ext in {".json"}:
            mimetype = "application/json; charset=utf-8"
        elif ext in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            mimetype = "image/" + ext[1:]
        with open(full_path, "rb") as fh:
            data = fh.read()
        self.send_response(200)
        self.send_header("Content-Type", mimetype)
        self.send_header("Content-Length", str(len(data)))
        # Uploads always land under a fresh unique filename, so item photos can be
        # cached hard. Without this the board re-fetches every photo on each
        # re-render and shows empty frames while they come back over the wire.
        if safe_path.startswith("images/"):
            self.send_header("Cache-Control", "public, max-age=604800, immutable")
        self.end_headers()
        self.wfile.write(data)
        return True

    def _handle_websocket(self):
        key = self.headers.get("Sec-WebSocket-Key")
        if not key:
            self.send_error(400)
            return
        self.protocol_version = "HTTP/1.1"
        self.send_response(101, "Switching Protocols")
        self.send_header("Upgrade", "websocket")
        self.send_header("Connection", "Upgrade")
        self.send_header("Sec-WebSocket-Accept", ws_accept_key(key))
        self.end_headers()
        sock = self.connection
        with ws_clients_lock:
            ws_clients.append(sock)
        try:
            while True:
                # We only push from server to client; any data/close/disconnect from
                # the client is treated the same way — stop serving this connection.
                data = sock.recv(6)
                if not data:
                    break
        except Exception:
            pass
        finally:
            with ws_clients_lock:
                if sock in ws_clients:
                    ws_clients.remove(sock)

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/ws" and self.headers.get("Upgrade", "").lower() == "websocket":
            self._handle_websocket()
            return
        if path == "/api/menu":
            try:
                self._send_json(read_menu_items(find_workbook_path()))
            except Exception as exc:
                self._send_json({"error": str(exc)}, 500)
            return
        if path == "/api/layout":
            try:
                if os.path.exists(LAYOUT_PATH):
                    with open(LAYOUT_PATH, "r", encoding="utf-8") as fh:
                        self._send_json(json.load(fh))
                else:
                    self._send_json({})
            except Exception as exc:
                self._send_json({"error": str(exc)}, 500)
            return
        self._serve_file(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/menu":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(content_length).decode("utf-8")
                payload = json.loads(body) if body else []
                if not isinstance(payload, list):
                    raise ValueError("Expected a JSON array")
                workbook_path = find_workbook_path()
                write_menu_items(workbook_path, payload)
                broadcast_ws(json.dumps({"type": "menu-updated"}))
                self._send_json({"ok": True, "path": os.path.basename(workbook_path)})
            except Exception as exc:
                self._send_json({"error": str(exc)}, 500)
            return
        if path == "/api/layout":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(content_length).decode("utf-8")
                payload = json.loads(body) if body else {}
                with open(LAYOUT_PATH, "w", encoding="utf-8") as fh:
                    json.dump(payload, fh, ensure_ascii=False)
                self._send_json({"ok": True})
            except Exception as exc:
                self._send_json({"error": str(exc)}, 500)
            return
        if path == "/api/upload-image":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(content_length).decode("utf-8")
                payload = json.loads(body) if body else {}
                data_url = payload.get("dataUrl", "")
                if not data_url.startswith("data:image/webp;base64,"):
                    raise ValueError("Expected a base64-encoded image/webp data URL")
                raw = base64.b64decode(data_url.split(",", 1)[1])
                os.makedirs(IMAGES_DIR, exist_ok=True)
                filename = unique_image_filename(slugify(payload.get("name", "item")))
                with open(os.path.join(IMAGES_DIR, filename), "wb") as fh:
                    fh.write(raw)
                self._send_json({"ok": True, "path": "images/" + filename})
            except Exception as exc:
                self._send_json({"error": str(exc)}, 500)
            return
        if path == "/api/shutdown":
            # Loopback-only: this is a launcher control signal, not a public API.
            if self.client_address[0] not in ("127.0.0.1", "::1"):
                self._send_json({"error": "forbidden"}, 403)
                return
            self._send_json({"ok": True})
            shutdown_event.set()
            return
        self._send_json({"error": "not found"}, 404)

    def log_message(self, format, *args):
        return


if __name__ == "__main__":
    host = os.environ.get("MENU_SERVER_HOST", "0.0.0.0")
    port = int(os.environ.get("MENU_SERVER_PORT", "8080"))
    try:
        server = ThreadingHTTPServer((host, port), MenuHandler)
    except OSError:
        print(f"Port {port} is already in use — a menu server is probably already running.")
        print(f"Reusing it: open http://localhost:{port}/admin.html or http://localhost:{port}/digital-menu.html")
    else:
        print(f"Menu admin:  http://localhost:{port}/admin.html")
        print(f"Menu board:  http://localhost:{port}/digital-menu.html")

        # serve_forever() runs on a background thread so the main thread is free to
        # block on shutdown_event — set either by POST /api/shutdown (the launcher's
        # graceful-stop request) or by Ctrl+C below. Calling server.shutdown() from
        # the same thread that's running serve_forever() would deadlock, which is
        # why they're split like this.
        server_thread = threading.Thread(target=server.serve_forever, daemon=True)
        server_thread.start()
        try:
            shutdown_event.wait()
        except KeyboardInterrupt:
            pass
        server.shutdown()
        server.server_close()
        server_thread.join(timeout=5)
        print("Server stopped.")
